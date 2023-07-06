from ai_auth.models import AiUser,UserCredits
from django.db.models import Q
from pathlib import Path
import json
import pandas as pd
import numpy as np
from collections import Counter,OrderedDict
from ai_workspace.models import Project,Job
from djstripe.models import Subscription,Charge
import logging
from django.db.models import Count
from django.utils import timezone
from datetime import timedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.core.mail import EmailMessage
import time
logger = logging.getLogger('django')
# FIXTURE_DIR_PATH = Path(__file__).parent.joinpath("fixtures")


# with FIXTURE_DIR_PATH.joinpath("test_users.json").open("r") as f:
#         test_users =json.load(f)



class AilaysaReport:
    output_xlsx = BytesIO()

    def get_users(self,is_vendor=False,period=None,test=False):
        if test:
            users= AiUser.objects.filter(~Q(email__icontains='deleted')&~Q(email='AnonymousUser')&~Q(is_staff=True) \
                    &~Q(is_internal_member=True)&Q(is_vendor=is_vendor))
        else:
            users= AiUser.objects.filter(~Q(email__icontains='deleted')&~Q(email='AnonymousUser')&~Q(is_staff=True) \
                    &~Q(is_internal_member=True)&Q(is_vendor=is_vendor)&~Q(email__icontains="+"))
        if period != None:
            start_date = timezone.now()
            end_date =start_date + timedelta(period)
            users = users.filter(date_joined__range=[start_date,end_date])
        return users

    def get_projects_and_langs(self,users):
        tot_langpairs,langpair=[],[]
        pairs=dict()
        for user in users:
            projs = Project.objects.filter(ai_user=user)
            for proj in projs:
                jobs = proj.project_jobs_set.all()
                for job in jobs:
                    #print(job.__str__())
                    langpair.append(job.__str__())
            pairs[user]=langpair
            tot_langpairs.extend(pairs)
        tot_counts = Counter(tot_langpairs)
       # return pairs,tot_langpairs,

    def project_count(self,users):
        projs_count=dict()
        for user in users:
            projs = Project.objects.filter(ai_user=user)
            projs_count[user]=projs.count()   
        return dict(sorted(projs_count.items(), key=lambda item: item[1],reverse=True))

    def user_credits(self,users):
        user_credits=dict()
        for user in users:
            uss = UserCredits.objects.filter(user=user)
            buy,left=0,0
            for us in uss:
                buy=buy+us.buyed_credits
                left=left+us.credits_left
            user_credits[user]={'buy':buy,'left':left}
        return user_credits
    
    def user_and_countries(self,users):
        return Counter(users.values_list('country__name',flat=True)).most_common()
    
    def no_of_active_user(self,users):
        return Subscription.objects.filter(status__in=['trialing','active','past_due']). \
            filter(Q(customer__subscriber__in=users)&~Q(plan__product__name="Pro - V")).values \
                ('customer__email').annotate(dcount =Count("customer__email")).count()
    
    def paid_users(self,users):
        return Charge.objects.filter(customer__subscriber__in=users,status="succeeded").values('customer__email').annotate \
            (dcount =Count("customer__email"))
    
    def users_trial_ending(self,users):
        start_date = timezone.now()
        end_date =start_date + timedelta(8)

        return Subscription.objects.filter(status__in=['trialing']).filter(Q(customer__subscriber__in=users)&~Q(plan__product__name="Pro - V")).filter \
            (trial_end__range=[start_date,end_date]).values("trial_end")
    
    def vendors_added_week(self,users):
        vendors = AiUser.objects.filter(~Q(email__icontains='deleted')&~Q(email='AnonymousUser')&~Q(is_staff=True)&~Q(is_internal_member=True)&Q(is_vendor=True))
        one_week_ago = timezone.now() - timedelta(days=7)
        return vendors.filter(date_joined__gte=one_week_ago).values('date_joined')

    def users_by_plan(self,users):
        return Subscription.objects.filter(status__in=['trialing','active','past_due']).filter(Q(customer__subscriber__in=users)&~Q(plan__product__name="Pro - V")).values \
            ('plan__product__name').annotate(dcount =Count("plan__product__name"))
    
    def langpairs_and_users(self,users):
        from collections import Counter
        tot_langpairs=[]
        pairs=dict()
        for user in users:
            langpair=[]
            projs = Project.objects.filter(ai_user=user)
            for proj in projs:
                jobs = proj.project_jobs_set.all()
                for job in jobs:
                    #print(job.__str__())
                    langpair.append(job.__str__())
            pairs[user.email]=langpair
        for key, value in pairs.items():
            tot_langpairs.extend(value)
        return Counter(tot_langpairs).most_common()
            
    def user_subscription_info(self,users,status=['trialing','active','past_due']):

        ## subscriptions data returns (sub_id,plan_name,plan_status,user_email)
        res = Subscription.objects.filter(status__in=status).filter \
        (Q(customer__subscriber__in=users)&~Q(plan__product__name="Pro - V")).values \
            ('customer__email').annotate(dcount =Count("customer__email")).filter(dcount=1)
        res_err = Subscription.objects.filter(status__in=status).filter \
        (Q(customer__subscriber__in=users)&~Q(plan__product__name="Pro - V")).values \
            ('customer__email').annotate(dcount =Count("customer__email")).filter(dcount__gte=2)
        if res_err.count()>0:
            logger.error("users have more than one active subscriptions")
        return res
    
    def user_subscription_plans(self,users):
        subs_info_active = Subscription.objects.filter(Q(customer__subscriber__in=users)&~Q(status__in=['active']
            )).values('plan__product__name').annotate(Count('plan__product__name'))
        subs_info_trial = Subscription.objects.filter(Q(customer__subscriber__in=users)&~Q(status__in=['trialing']
            )).values('plan__product__name').annotate(Count('plan__product__name'))
        return subs_info_active,subs_info_trial
    
    def total_languages_used(self,users=None):
        jobs1 = Job.objects.filter(source_language__isnull=False).values('target_language__language'
                ).distinct().values_list('source_language__language',flat=True)
        jobs2 = Job.objects.filter(target_language__isnull=False).values('target_language__language'
                ).distinct().values_list('target_language__language',flat=True)

        job1 = list(jobs1)
        job2 = list(jobs2)
        return set(job1+job2)


    def users_details(self,users,proj_details,user_credits_ls,subs_details):
        
        data =[]
        df = pd.DataFrame(data, columns=['UID',
            'Email',
            'Country',
            'Projects Created',
            #'Langpair Used',
            'Intial Credits',
            'Credits Left',
            'Previous plan',
            'Previous plan status'
            'Current Plan',
            'Current Plan status'])
        for user in users:

            if proj_details:
                proj_count = proj_details.get(user)
            
            if user_credits_ls:
                user_credi_details = user_credits_ls.get(user)
                print(user_credi_details)


            if subs_details:
                try:
                    subs_details.get(customer__email=user.email).get('customer__email')
                    sub = Subscription.objects.get(customer__email=user.email,status__in=['trialing','active','past_due'])
                    sub = Subscription.objects.filter(Q(customer__email=user.email)&~Q(status__in=['trialing','active','past_due']))
                    plan_name = sub.plan.product.name
                    status = sub.status
                except:
                    sub = None
                    plan_name = None
                    status = None
                
            if user.country == None:
                country_name = None
            else:
                country_name = user.country.name

            df2 = {'UID':user.uid,'Email':user.email, 'Country': country_name, 'Projects Created': proj_count,'Intial Credits':user_credi_details.get('buy'),'Credits Left':user_credi_details.get('left'),
                  'Current Plan':plan_name,'Current Plan status':status,'Created':pd.to_datetime(user.date_joined.replace(tzinfo=None))}
            df = df.append(df2, ignore_index = True)
            # df.insert(loc=0, column='UID', value=player_vals)
            # df.insert(loc=0, column='Email', value=player_vals)
            # df.insert(loc=0, column='Country', value=player_vals)
        return df
        
        # df = pd.DataFrame.from_dict({
        #     'UID': ['Nik', 'Kate', 'Evan', 'Kyra'],
        #     'Email': [31, 30, 40, 33],
        #     'Country': ['Toronto', 'London', 'Kingston', 'Hamilton'],
        #     'Projects Created':[0,1,2,4],
        #     'Langpair Used':['English-Tamil','Tamil-French','English-Tamil','English-Hindhi'],
        #     'Intial Credits':[2000,2000,2000,2000],
        #     'Credits Left':[100,200,202,20],
        #     'Plan':['pro','pro','Business','pro'],
        #     'Subscription status':['active','inactive','trialing','past_due']
        # })

    #df = df.append({'Name':'Jane', 'Age':25, 'Location':'Madrid'}, ignore_index=True)

       # df.to_excel('ailaysa-report.xlsx',index=False)
        # writer = pd.ExcelWriter("SalesReport.xlsx")
        # repr_sales.to_excel(writer, sheet_name="Sale per rep")

    def users_count_monthly(self,df):
        df3 = df.groupby(pd.Grouper(key='Created', axis=0, freq='M')).count()
        df4 = pd.DataFrame({'Created at':[cel.date().strftime("%Y-%b") for cel in df3.index],
                    'No of users':[mail for mail in df3.Email]})
        print("df4",df4)
        return df4

    # def chart_gen(self,df):
    #    df3 = df.groupby(pd.Grouper(key='Created', axis=0, freq='M')).count()
    #    df3[df3['Created at'].dt.date.astype(str) == '2022-12-12']

    def users_stats(self,users):
        subs_details_trial = self.user_subscription_info(users,status=['trialing',])
        subs_details_past_due = self.user_subscription_info(users,status=['past_due',])
        users_week = self.get_users(period=8)
        vendors_week = self.get_users(is_vendor=True,period=8)
        paid_users =self.paid_users(users)
        df = pd.DataFrame(data =([subs.get('customer__email') for subs in subs_details_trial]),columns =['user in trial'])
        df1 = pd.DataFrame(data =([subs.get('customer__email') for subs in subs_details_past_due]),columns =['user in past_due'])
        df2 = pd.DataFrame(data =([user.email for user in users_week]),columns =['users added in week'])
        df3 = pd.DataFrame(data =([vendor.email for vendor in vendors_week]),columns =['Vendor added in week'])
        df4 = pd.DataFrame(data =([user.get('customer__email') for user in paid_users]),columns =['paid users'])
        result = pd.concat([df,df1,df2,df3,df4],axis=1)
        return result

    def color_header(self,ws):
        for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
            for cell in rows:
                cell.fill = PatternFill(bgColor="00FFFF00")

    def create_chart(self,ws1,ws2):
        dim = self.get_len_sheet(ws1)
        values = Reference(ws1,
                        min_col=2,
                        max_col=2,
                        min_row=2,
                        max_row=dim['B'])
        cats = Reference(ws1, min_col=1, max_col=1, min_row=2, max_row=dim['A'])

        # Create object of BarChart class
        chart = BarChart()
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(cats)

        # set the title of the chart
        chart.title = "Users Count"

        # set the title of the x-axis
        chart.x_axis.title = "Created date"

        # set the title of the y-axis
        chart.y_axis.title = "no of users"

        # the top-left corner of the chart
        # is anchored to cell F2 .
        ws2.add_chart(chart,"F2")


    def create_chart_countries(self,ws1,ws2):
        dim = self.get_len_sheet(ws1)
        values = Reference(ws1,
                        min_col=2,
                        max_col=2,
                        min_row=2,
                        max_row=dim['B'])
        cats = Reference(ws1, min_col=1, max_col=1, min_row=2, max_row=dim['A'])

        # Create object of BarChart class
        chart = BarChart()
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(cats)

        # set the title of the chart
        chart.title = "Ailaysa Signup Countries"

        # set the title of the x-axis
        chart.x_axis.title = "Countries"

        # set the title of the y-axis
        chart.y_axis.title = "No Of Users"

        # the top-left corner of the chart
        # is anchored to cell F2 .
        ws2.add_chart(chart,"F21")

    def create_excel(self,users):
        wb = Workbook()
        ws1 = wb.create_sheet("Language Pairs")
        ws2 = wb.create_sheet("Countries")
        ws3 = wb.create_sheet("Projects")
        ws4 = wb.create_sheet("Users List")
        ws5 = wb.create_sheet("Users Details")
        ws6 = wb.create_sheet("Users Signup")
        ws7 = wb.create_sheet("Charts")
        tot_pairs = self.langpairs_and_users(users)
        user_countries = self.user_and_countries(users)
        proj_count =self.project_count(users)
        projs_det = self.project_count(users)
        user_credits_ls = self.user_credits(users)
        subs_details = self.user_subscription_info(users)
        

        df4 = pd.DataFrame({'Language pairs':[pair[0] for pair in tot_pairs],'No of Task':[pair[1] for pair in tot_pairs]})
        for r in dataframe_to_rows(df4, index=False, header=True):
            ws1.append(r)
        df5 = pd.DataFrame({'Countries':[country[0] for country in user_countries],'No of users':[country[1] for country in user_countries]})
        for r in dataframe_to_rows(df5, index=False, header=True):
            ws2.append(r)
        df6 = pd.DataFrame({'Users':[key.email for key in proj_count.keys()],'No of projects':[value for value in proj_count.values()]})
        for r in dataframe_to_rows(df6, index=False, header=True):
            ws3.append(r)
        df7 = self.users_stats(users)
        for r in dataframe_to_rows(df7, index=False, header=True):
            ws4.append(r)
        df8 = self.users_details(users,projs_det,user_credits_ls,subs_details)
        for r in dataframe_to_rows(df8, index=False, header=True):
            ws5.append(r)
        df9 = self.users_count_monthly(df8)
        for r in dataframe_to_rows(df9, index=False, header=True):
            ws6.append(r)
        
        self.create_chart(ws6,ws7)
        self.create_chart_countries(ws2,ws7)

        # ws1=self.color_header(ws1)
        # ws2=self.color_header(ws2)
        # ws3=self.color_header(ws3)
        # ws4=self.color_header(ws4)
        # ws5=self.color_header(ws5)

        wb.save("ai_reports.xlsx")
        return wb

    def report_generate(self):
        users = self.get_users()
        wb = self.create_excel(users.order_by('-date_joined'))
        wb.save(self.output_xlsx)

    def get_len_sheet(self,ws):
        dimensions = dict()
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            max_col_row = len([cell for cell in ws[col_letter] if cell.value != ''])
            dimensions[col_letter] =max_col_row
        return dimensions

    def send_report(self):
        filename = f'''ailaysa_report_{time.strftime("%Y%m%d-%H%M%S")}.xlsx'''

        email = EmailMessage(
        'Ailaysa Reports',
        'Ailaysa Report Generated In Production',
        'noreply@ailaysa.com',
        ['stephenlangtest@gmail.com'],
        #reply_to=['another@example.com'],
        #headers={'Message-ID': 'foo'},
        )
        email.attach(filename, self.output_xlsx.getvalue() , 'application/vnd.ms-excel')
        if email.send():
            logger.info("Report sent")
        else:
            logger.error("Issue in report sending")